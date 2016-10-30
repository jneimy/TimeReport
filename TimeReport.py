from __future__ import division
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import date, timedelta, datetime
import os
import sys
import requests
import json
import re
import argparse

excel_filename = "Timesheet_Report.xlsx"

harvest_headers = {
	'Content-type': 'application/json',
	'Accept': 'application/json',
	'Authorization': 'Basic Y21vcmlrdW5pQHJldmFjb21tLmNvbToqNjAlaEZ4ViVSWHU='
}

hoursForAcceptance = 5

# Excel Style
namesStart = (4, 1)
totCell = (1, 7)
consultantsFont = Font(name='Calibri',
							size=12,
							color='00FFFFFF')
consultantsFill = PatternFill(start_color='001F4E78',
								end_color='001F4E78',
								fill_type='solid')
totalsFont = Font(name='Calibri',
					size=12,
					bold=True)
text = Font(name='Calibri',
					size=12)

def init():
	# Harvest - Build request & load projects
	projects = requests.get('https://revacomm.harvestapp.com/projects', headers=harvest_headers)
	projects_json = projects.json()

	return projects_json


def peopleTime(date):
	fteTime = {}
	contTime = {}
	people = requests.get('https://revacomm.harvestapp.com/people', headers=harvest_headers)
	people_json = people.json()
	for person in people_json:
		pUser = person['user']
		uid = str(pUser['id'])
		first = pUser['first_name']
		last = pUser['last_name']

		# Skip old employees
		if not pUser['is_active']:
			continue

		# Identify contractors
		isContract = pUser['is_contractor']

		userTime = requests.get('https://revacomm.harvestapp.com/people/' + uid + '/entries?from=' + yesterday + '&to=' + yesterday, headers=harvest_headers)
		userTime_json = userTime.json()

		marked = False
		enteredTime = 0
		if userTime_json is not None:
			for ut in userTime_json:
				entry = ut.get("day_entry", 0)
				if entry == 0:
					continue
				hours = entry.get("hours", 0)
				enteredTime += hours
			if enteredTime >= hoursForAcceptance:
				marked = True
			# Check if intern argument is called
				# If called, check whether current person is an intern
				# If person is an intern, check if they work during dayOfWeek
				# If they work, check if enteredTime > 0
				# If 0, marked = True
				# test
		if isContract:
			contTime[first + " " + last] = marked
		else:
			fteTime[first + " " + last] = marked
	return (fteTime, contTime)


def openExcel(filename, weekSheetName, fteTime, contTime):
	wb = None
	ws = None

	# Open or create a new worksheet
	isCreateWs = False
	if os.path.exists(filename):
		wb = load_workbook(filename)
		if weekSheetName in wb.get_sheet_names():
			ws = wb[weekSheetName]
		else:
			isCreateWs = True
	else:
		wb = Workbook()
		ws = wb.active

		# Delete current sheet & create new
		if ws is not None:
			wb.remove_sheet(ws)
		isCreateWs = True

	if isCreateWs:
		wsSrc = wb["Template"]
		ws = wb.copy_worksheet(wsSrc)
		ws.title = weekSheetName
		ws.cell(row=1, column=1, value=weekSheetName)

		# Reorder
		sheetInd = len(wb.get_sheet_names()) - 1
		wb._sheets = [wb._sheets[sheetInd]] + wb._sheets[0:sheetInd]
		ws.active = 0

		# Setup Template
		row = namesStart[0]
		for key in sorted(fteTime.iterkeys()):
			ws.cell(row=row, column=1, value=key)
			row += 1

		# Consultant Header
		row += 1
		cell = ws.cell(row=row, column=1, value="Consultants")
		cell.fill = consultantsFill
		cell.font = consultantsFont
		row += 1

		for key in sorted(contTime.iterkeys()):
			ws.cell(row=row, column=1, value=key)
			row += 1

		# Setup Formulas
		row += 1
		cell = ws.cell(row=row, column=1, value="Totals")
		cell.font = totalsFont

		formRowStart = 4
		formRowEnd = row - 1
		formColStart = 'B'
		formColEnd = 'G'

		# Total Formula
		totArea = formColStart + str(formRowStart) + ':' + formColEnd + str(formRowEnd)
		ws[formColEnd + '1'] = "=SUM(" + totArea + ")/COUNT(" + totArea + ")"

		# Day Total Formula
		for ordCol in range(ord(formColStart), ord(formColEnd)+1):
			dayTotArea = chr(ordCol) + str(formRowStart) + ':' + chr(ordCol) + str(formRowEnd)
			ws[chr(ordCol) + str(row)] = "=SUM(" + dayTotArea + ")/COUNT(" + dayTotArea + ")"
			ws[chr(ordCol) + str(row)].number_format = '0%'
	return (wb, ws)


def closeExcel(wb, filename):
	wb.save(filename)


def outputToExcel(ws, dayOfWeek, fteTime, contTime, ignoreList):
	# 2 for one space and base 1
	dayToCol = dayOfWeek + 2

	row = namesStart[0]
	for key in sorted(fteTime.iterkeys()):
		if key.lower() in ignoreList:
			row += 1
			continue
		cell = ws.cell(row=row, column=1)
		# Skip FTE weekends
		if dayOfWeek <= 4:
			if key == cell.value:
				ws.cell(row=row, column=dayToCol, value=int(fteTime[key]))
			else:
				print "ERROR: invalid person key: " + key + " cell: " + cell.value
				continue
		row += 1

	# increase row to start of contractors
	row += 2
	for key in sorted(contTime.iterkeys()):
		if key.lower() in ignoreList:
			row += 1
			continue
		cell = ws.cell(row=row, column=1)
		# Skip contractor weekends Monday & Sunday
		if dayOfWeek != 0 and dayOfWeek != 6:
			if key == cell.value:
				ws.cell(row=row, column=dayToCol, value=int(contTime[key]))
			else:
				print "ERROR: invalid person key: " + key + " cell: " + cell.value
				continue
		row += 1


if __name__ == '__main__':
	## CI runs this at 8am HNL ##
	parser = argparse.ArgumentParser(description="Calculate time reports.")
	parser.add_argument("--date", help="mm.dd.yyyy of date to gather")
	parser.add_argument("--ignore", help="comma seperated list of people to ignore")
	parser.add_argument("--interns", help="Use JSON format as arguments for the intern's schedule. Example: {\"Aljon Preza\": [0,1,2]}")
	args = parser.parse_args()

        internArgs = None;
	if args.interns:
		internArgs = json.loads(args.interns)
		for internMember in internArgs:
                        print internMember # Gets intern name
                        print internArgs.get(internMember) # Gets the value of the intern (work days)
                        internWorkDays = internArgs.get(internMember); # Store work day array into internWorkDays
                        print "I work on %d" % (internWorkDays[0]) # Gets the first day that the intern works
                        
                internTest = "Aljon Preza" # Intern to test
                if internArgs.get(internTest): # Checks whether the intern is part of the dict
                        print "\n%s is an intern and works in the following hours:" % internTest # Prints intern
                        print internArgs.get(internTest)
                sys.exit(0)
                

	setDt = None
	if args.date:
		setDt = datetime.strptime(args.date, "%m.%d.%Y")

	ignoreList = []
	if args.ignore:
		ignoreList = [x.strip().lower() for x in args.ignore.split(',')]
		print "IGNORE: " + ", ".join(ignoreList).title()

	if setDt:
		yesterdayDt = setDt
	else:
		yesterdayDt = date.today() - timedelta(1)
	yesterday = str(yesterdayDt.strftime('%Y%m%d'))
	yesterdayFmt = str(yesterdayDt.strftime('%m.%d.%Y'))
	print "Time Reporting for: " + yesterdayFmt

	# # M=0, T=1, W=2, T=3, F=4, S=5, S=6
	dayOfWeek = yesterdayDt.weekday()
	startOfWeekDt = yesterdayDt - timedelta(dayOfWeek)
	startOfWeekFmt = str(startOfWeekDt.strftime('%m.%d.%Y'))

	(fteTime, contTime) = peopleTime(yesterday)

	wb, ws = openExcel(excel_filename, startOfWeekFmt, fteTime, contTime)
	outputToExcel(ws, dayOfWeek, fteTime, contTime, ignoreList)
	closeExcel(wb, excel_filename)
